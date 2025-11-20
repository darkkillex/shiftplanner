import calendar, logging, re
from collections import defaultdict

from io import BytesIO

from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response

from email.utils import formataddr

from django.template.loader import render_to_string
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.db import transaction
from django.db.models import F, Subquery, Count, Q
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import logout
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from .forms import PlanCreateForm
from .forms import TemplateCreateForm
from .models import Template, TemplateRow, Profession, AssignmentSnapshot, PlanRow, Employee, Company, Assignment

from .serializers import *

_SLOT_RE = re.compile(r'^(.*?)(?:\.(\d+))?$')

def _split_slot(name: str):
    m = _SLOT_RE.match((name or '').strip())
    base = (m.group(1) or '').strip()
    num = int(m.group(2) or 0)
    return base, num

def _existing_max_suffix(base: str) -> int:
    # Rileva il massimo suffisso esistente per quella base
    names = Profession.objects.filter(name__regex=rf'^{re.escape(base)}(?:\.(\d+))?$').values_list('name', flat=True)
    m = 0
    for n in names:
        _, k = _split_slot(n)
        if k > m:
            m = k
    return m

def _assign_signature(a):
    return f"{a.shift_type_id or ''}|{a.profession_id or ''}|{(a.notes or '').strip()}"

def _next_suffix_for_base(base: str) -> int:
    # guarda tutte le Profession già esistenti per quella base
    names = Profession.objects.filter(
        name__regex=rf'^{re.escape(base)}(?:\.(\d+))?$'
    ).values_list('name', flat=True)
    mx = 0
    for n in names:
        _, k = _split_slot(n)
        if k > mx: mx = k
    return mx + 1

def _ensure_profession_for(base: str, explicit_num: int | None) -> Profession:
    if explicit_num and explicit_num > 0:
        name = f"{base}.{explicit_num}"
    else:
        name = f"{base}.{_next_suffix_for_base(base)}"
    prof, _ = Profession.objects.get_or_create(name=name)
    return prof

def _propagate_insert_to_plans(template: Template, insert_order: int, duty: str):
    BUMP = 1000
    plans = Plan.objects.filter(template=template).prefetch_related('rows')
    for plan in plans:
        # se già esiste una riga con quel duty, salta
        if plan.rows.filter(duty=duty).exists():
            continue
        # bump per evitare collisione su uniq_plan_order
        PlanRow.objects.filter(plan=plan, order__gte=insert_order).update(order=F('order') + BUMP)
        PlanRow.objects.create(plan=plan, order=insert_order, duty=duty, is_spacer=False)
        _normalize_plan_orders(plan)

def _normalize_template_orders(template: Template):
    rows = list(TemplateRow.objects.filter(template=template).order_by('order', 'id'))
    for i, r in enumerate(rows, start=1):
        if r.order != i:
            r.order = i
            r.save(update_fields=['order'])

def _normalize_plan_orders(plan):
    rows = list(PlanRow.objects.filter(plan=plan).order_by('order', 'id'))
    for i, r in enumerate(rows, start=1):
        if r.order != i:
            r.order = i
            r.save(update_fields=['order'])


def _easter_sunday(year: int) -> dt.date:
    # algoritmo gregoriano (Anonymous Gregorian)
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19*a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2*e + 2*i - h - k) % 7
    m = (a + 11*h + 22*l) // 451
    month = (h + l - 7*m + 114) // 31
    day = 1 + ((h + l - 7*m + 114) % 31)
    return dt.date(year, month, day)

def italy_holidays(year: int) -> set[dt.date]:
    """Festività nazionali italiane principali + Lunedì dell’Angelo."""
    easter = _easter_sunday(year)
    pasquetta = easter + dt.timedelta(days=1)

    fixed = {
        (1, 1),   # Capodanno
        (1, 6),   # Epifania
        (4, 25),  # Liberazione
        (5, 1),   # Lavoro
        (6, 2),   # Repubblica
        (8, 15),  # Assunzione
        (11, 1),  # Ognissanti
        (12, 8),  # Immacolata
        (12, 25), # Natale
        (12, 26), # Santo Stefano
    }
    hs = {dt.date(year, m, d) for (m, d) in fixed}
    hs.add(pasquetta)
    return hs

def _range_from_preset(preset: str) -> tuple[dt.date, dt.date]:
    today = dt.date.today()
    if preset == "m1":   # ultimo mese
        start = (today.replace(day=1) - dt.timedelta(days=1)).replace(day=1)
        end_last = (today.replace(day=1) - dt.timedelta(days=1))
        return start, end_last
    if preset == "q1":   # ultimo trimestre
        m = ((today.month - 1)//3)*3 + 1
        q_start = dt.date(today.year, m, 1)
        prev_q_end = q_start - dt.timedelta(days=1)
        prev_m = ((prev_q_end.month - 1)//3)*3 + 1
        prev_q_start = dt.date(prev_q_end.year, prev_m, 1)
        return prev_q_start, prev_q_end
    if preset == "h1":   # ultimo semestre
        m = 1 if today.month <= 6 else 7
        sem_start = dt.date(today.year, m, 1)
        prev_sem_end = sem_start - dt.timedelta(days=1)
        prev_m = 1 if prev_sem_end.month <= 6 else 7
        prev_sem_start = dt.date(prev_sem_end.year, prev_m, 1)
        return prev_sem_start, prev_sem_end
    if preset == "y1":   # ultimo anno solare completo
        start = dt.date(today.year-1, 1, 1)
        end   = dt.date(today.year-1, 12, 31)
        return start, end
    # default: ultimi 30 giorni rolling
    return today - dt.timedelta(days=29), today

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all().order_by('last_name','first_name')
    serializer_class = EmployeeSerializer

class ProfessionViewSet(viewsets.ModelViewSet):
    queryset = Profession.objects.all().order_by('name')
    serializer_class = ProfessionSerializer

class ShiftTypeViewSet(viewsets.ModelViewSet):
    queryset = ShiftType.objects.all().order_by('id')
    serializer_class = ShiftTypeSerializer

class ReminderViewSet(viewsets.ModelViewSet):
    serializer_class = ReminderSerializer
    queryset = Reminder.objects.all().order_by("date","completed","title")

    def get_queryset(self):
        qs = super().get_queryset()
        # filtra per mese: ?year=2025&month=10
        y = self.request.query_params.get("year")
        m = self.request.query_params.get("month")
        if y and m:
            from calendar import monthrange
            y, m = int(y), int(m)
            start = dt.date(y, m, 1)
            end = dt.date(y, m, monthrange(y, m)[1])
            qs = qs.filter(date__range=(start, end))
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

class PlanViewSet(viewsets.ModelViewSet):
    queryset = Plan.objects.all().order_by('-year','-month')
    serializer_class = PlanSerializer

    @action(detail=True, methods=['get'])
    def grid(self, request, pk=None):
        plan = self.get_object()
        days = calendar.monthrange(plan.year, plan.month)[1]

        ass = (Assignment.objects
               .filter(plan=plan)
               .select_related('employee', 'shift_type', 'profession'))
        idx = {(a.profession_id, a.date): a for a in ass}

        rows = []
        if plan.rows.exists():
            for r in plan.rows.all().order_by('order'):
                if r.is_spacer:
                    duty_full = ''
                    base = ''
                    prof = None
                else:
                    duty_full = (r.duty or '').strip()  # es. "Magazzino.4"
                    base = duty_full.split('.', 1)[0] if duty_full else ''
                    prof = Profession.objects.filter(name=duty_full).first()  # match ESATTO

                row = {
                    'plan_row_id': r.id,
                    'profession_id': getattr(prof, 'id', None),
                    'profession': base,  # UI senza suffisso
                    'spacer': bool(r.is_spacer),
                }

                for d in range(1, days + 1):
                    day = dt.date(plan.year, plan.month, d)
                    if not prof:
                        row[str(d)] = {'employee_id': None, 'employee_name': '', 'shift_code': '', 'shift_label': '',
                                       'notes': '', 'has_note': False}
                    else:
                        a = idx.get((prof.id, day))
                        row[str(d)] = {
                            'employee_id': a.employee_id if a else None,
                            'employee_name': a.employee.full_name() if a else '',
                            'shift_code': a.shift_type.code if (a and a.shift_type) else '',
                            'shift_label': a.shift_type.label if (a and a.shift_type) else '',
                            'notes': a.notes if a and a.notes else '',
                            'has_note': bool(a and a.notes),
                        }
                rows.append(row)
        else:
            for p in Profession.objects.order_by('name'):
                base = p.name.split('.', 1)[0]
                row = {'profession_id': p.id, 'profession': base, 'spacer': False}
                for d in range(1, days + 1):
                    day = dt.date(plan.year, plan.month, d)
                    a = idx.get((p.id, day))
                    row[str(d)] = {
                        'employee_id': a.employee_id if a else None,
                        'employee_name': a.employee.full_name() if a else '',
                        'shift_code': a.shift_type.code if (a and a.shift_type) else '',
                        'shift_label': a.shift_type.label if (a and a.shift_type) else '',
                        'notes': a.notes if a and a.notes else '',
                        'has_note': bool(a and a.notes)
                    }
                rows.append(row)

        return Response({'year': plan.year, 'month': plan.month, 'days': days, 'rows': rows})

    @action(detail=True, methods=['post'])
    def bulk_assign(self, request, pk=None):
        plan = self.get_object()
        employee_id = request.data.get('employee_id')
        shift_id = request.data.get('shift_type_id')
        cells = request.data.get('cells', [])
        note = (request.data.get('note') or '').strip()

        if not employee_id or not isinstance(cells, list) or not cells:
            return Response({'detail': 'employee_id e cells obbligatori'}, status=400)

        # mappa professioni per nome (tollerante)
        prof_by_name = {p.name.strip().casefold(): p for p in Profession.objects.all()}

        valid_targets = []  # [(profession_id:int, date:date)]
        skipped = []  # [{'date':iso, 'reason': str}]
        errors = []  # hard errors di formato

        # normalizza celle
        for c in cells:
            try:
                d_iso = str(c.get('date'))
                day = dt.date.fromisoformat(d_iso)
            except Exception:
                errors.append({'cell': c, 'reason': 'date non valida'})
                continue

            prof_id = c.get('profession_id')
            if prof_id:
                # verifica che la professione esista
                if Profession.objects.filter(pk=prof_id).exists():
                    valid_targets.append((int(prof_id), day))
                    continue
                else:
                    skipped.append({'date': d_iso, 'reason': 'profession_id inesistente'})
                    continue

            # fallback: plan_row_id -> mappa duty->Profession
            pr_id = c.get('plan_row_id')
            if pr_id:
                from .models import PlanRow  # import locale, nessuna hard dep se il modello manca
                try:
                    r = PlanRow.objects.select_related(None).get(pk=pr_id, plan=plan)
                    if r.is_spacer or not (r.duty or '').strip():
                        skipped.append({'date': d_iso, 'reason': 'riga spacer/non mappabile'})
                        continue
                    key = r.duty.strip().casefold()
                    p = prof_by_name.get(key)
                    if not p:
                        skipped.append({'date': d_iso, 'reason': f"mansione '{r.duty}' non mappata a Profession"})
                        continue
                    valid_targets.append((p.id, day))
                    continue
                except Exception:
                    skipped.append({'date': d_iso, 'reason': 'plan_row inesistente o di altro piano'})
                    continue

            # nessun id utile
            skipped.append({'date': d_iso, 'reason': 'nessun profession_id/plan_row_id'})
            continue

        if errors:
            # formato celle non valido
            return Response(
                {'detail': 'formato celle non valido', 'errors': errors, 'skipped': []},
                status=400
            )

        if not valid_targets:
            # nessuna cella mappabile
            return Response(
                {'detail': 'nessuna cella valida',
                 'skipped': skipped,  # <-- motivi per ogni data
                 'hint': 'verifica plan_row_id/profession_id e corrispondenza mansione->Profession'},
                status=400
            )
        log = logging.getLogger(__name__)
        log.info("bulk_assign plan=%s employee=%s cells=%s", plan.id, employee_id, cells[:5])
        # conflitti: stesso dipendente già assegnato in quelle date su ALTRE professioni
        dates = list({d for _, d in valid_targets})
        conflicts_qs = (Assignment.objects
                        .filter(plan=plan, date__in=dates, employee_id=employee_id)
                        .select_related('profession', 'shift_type'))
        target_set = set(valid_targets)  # confronto su (profession_id, date)

        conflicts = []
        for a in conflicts_qs:
            if (a.profession_id, a.date) in target_set:
                continue  # stessa cella: ok
            conflicts.append({
                'date': a.date.isoformat(),
                'profession_id': a.profession_id,
                'profession': a.profession.name if a.profession_id else '',
                'shift_code': a.shift_type.code if a.shift_type else '',
                'shift_label': a.shift_type.label if a.shift_type else '',
            })
        if conflicts:
            return Response(
                {'detail': 'conflitto assegnazioni sulle date selezionate', 'conflicts': conflicts},
                status=status.HTTP_409_CONFLICT
            )

        # write
        updated = 0
        with transaction.atomic():
            for prof_id, day in valid_targets:
                obj, _ = Assignment.objects.update_or_create(
                    plan=plan, profession_id=prof_id, date=day,
                    defaults={'employee_id': employee_id, 'shift_type_id': shift_id}
                )
                if obj.notes != note:
                    obj.notes = note
                    obj.save(update_fields=['notes'])
                updated += 1

        return Response({'updated': updated, 'skipped': skipped}, status=200)

    @action(detail=True, methods=['post'])
    @transaction.atomic
    def notify(self, request, pk=None):
        plan = self.get_object()
        logger = logging.getLogger(__name__)

        # --- Totali "griglia" sempre calcolati ---
        total_employees_registered = Employee.objects.count()

        emp_ids_qs = (Assignment.objects
                      .filter(plan=plan)
                      .values_list('employee_id', flat=True)
                      .distinct())
        total_in_plan = emp_ids_qs.count()
        with_email = (Employee.objects
                      .filter(id__in=Subquery(emp_ids_qs), email__isnull=False)
                      .exclude(email='')
                      .count())

        # --- Assegnazioni correnti con email valida (per invii reali) ---
        cur_qs = (Assignment.objects
                  .filter(plan=plan, employee__email__isnull=False)
                  .exclude(employee__email='')
                  .select_related('employee', 'shift_type', 'profession'))

        # Firma corrente: (emp_id, date) -> signature
        cur_sig = {}
        for a in cur_qs:
            cur_sig[(a.employee_id, a.date)] = _assign_signature(a)  # funzione esistente nel tuo codice

        # --- Snapshot mese ---
        snap_qs = AssignmentSnapshot.objects.filter(year=plan.year, month=plan.month)
        has_snapshot = snap_qs.exists()

        changed_emp_ids = set()
        if not has_snapshot:
            # Primo invio del mese: tutti quelli presenti con email valida
            changed_emp_ids = {a.employee_id for a in cur_qs}
        else:
            snap_sig = {(s.employee_id, s.date): s.signature for s in snap_qs}
            keys = set(cur_sig.keys()) | set(snap_sig.keys())
            # Un employee è "cambiato" se almeno una cella differisce
            for (eid, day) in keys:
                if cur_sig.get((eid, day)) != snap_sig.get((eid, day)):
                    # Notifica solo se l'employee è presente oggi nel piano (con email valida)
                    if any(k[0] == eid for k in cur_sig.keys()):
                        changed_emp_ids.add(eid)

        # Nessun destinatario: rispondi comunque coi totali
        if not changed_emp_ids:
            return Response({
                "total_employees_registered": total_employees_registered,
                "total_in_plan": total_in_plan,
                "with_email": with_email,
                "prepared": 0,
                "sent": 0,
                "recipients": [],
                "detail": "Nessuna variazione rispetto all’ultimo invio."
            }, status=200)

        # --- Grouping per destinatario, mantenendo i tuoi template ---
        by_emp = defaultdict(list)
        for a in cur_qs.filter(employee_id__in=changed_emp_ids):
            by_emp[a.employee].append(a)


        ITALIAN_MONTHS = [
            "", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
            "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"
        ]
        month_name = ITALIAN_MONTHS[plan.month]
        month_number = f"{plan.month:02d}"
        legend = list(ShiftType.objects.order_by('id').values_list('code', 'label'))
        sender_name = request.user.get_full_name() or request.user.username
        from_email = settings.DEFAULT_FROM_EMAIL or settings.EMAIL_HOST_USER
        reply_to = [settings.REPLY_TO_EMAIL] if getattr(settings, 'REPLY_TO_EMAIL', None) else None
        rev_str = ""  # lasciato vuoto come richiesto

        prepared = len(by_emp)
        sent = 0

        for emp, items in by_emp.items():
            items.sort(key=lambda x: x.date)
            rows = []
            for a in items:
                full_name = a.profession.name if a.profession else ''
                base_name, _ = _split_slot(full_name)
                rows.append({
                    'weekday': ['Lun', 'Mar', 'Mer', 'Gio', 'Ven', 'Sab', 'Dom'][a.date.weekday()],
                    'date': a.date.strftime('%d/%m/%Y'),
                    'profession_name': base_name,
                    'shift_label': a.shift_type.label if a.shift_type else '',
                    'notes': a.notes or '',
                })

            ctx = {
                'employee_name': emp.full_name(),
                # 'month_number': month_number,
                'month_name': month_name,
                'year': plan.year,
                'legend': legend,
                'assignments': rows,
                'app_url': f"{settings.APP_BASE_URL}/plan/{plan.pk}/",
                'reply_to_email': getattr(settings, 'REPLY_TO_EMAIL', from_email),
                'revision_str': rev_str,
                'sender_name': sender_name,
            }

            subject = f"Piano turni {month_name} {plan.year} {rev_str}".strip()
            text_body = render_to_string("emails/plan_personal.txt", ctx)
            html_body = render_to_string("emails/plan_personal.html", ctx)

            msg = EmailMultiAlternatives(
                subject=subject,
                body=text_body,
                from_email=from_email,
                to=[formataddr((emp.full_name(), emp.email))],
                reply_to=reply_to
            )
            msg.attach_alternative(html_body, "text/html")

            try:
                n = msg.send(fail_silently=False)
                sent += (1 if n else 0)
            except Exception as e:
                logger.exception("Errore invio a %s <%s>: %s", emp.full_name(), emp.email, e)

        # --- Aggiorna snapshot SOLO per gli employee notificati ---
        if prepared:
            AssignmentSnapshot.objects.filter(
                year=plan.year, month=plan.month, employee_id__in=changed_emp_ids
            ).delete()
            snaps = []
            for (eid, day), sig in cur_sig.items():
                if eid in changed_emp_ids:
                    snaps.append(AssignmentSnapshot(
                        year=plan.year, month=plan.month, employee_id=eid, date=day, signature=sig
                    ))
            if snaps:
                AssignmentSnapshot.objects.bulk_create(snaps, ignore_conflicts=True)

        recipients_list = [e.full_name() for e in by_emp.keys()]
        return Response({
            "total_employees_registered": total_employees_registered, # totale dipendenti registrati su Kairos
            "total_in_plan": total_in_plan,  # distinti presenti nel piano
            "with_email": with_email,  # di quelli, con email valida
            "prepared": prepared,  # cambiati e notificabili ora
            "sent": sent,  # invii effettivi
            "recipients": recipients_list  # nomi destinatari
        }, status=200)

    @action(detail=True, methods=['post'])
    def bulk_clear(self, request, pk=None):
        plan = self.get_object()
        cells = request.data.get('cells', [])
        if not isinstance(cells, list) or not cells:
            return Response({'detail': 'cells obbligatorio'}, status=status.HTTP_400_BAD_REQUEST)

        deleted = 0
        for c in cells:
            try:
                prof_id = int(c['profession_id'])
                day = dt.date.fromisoformat(c['date'])  # forza tipo date, niente ambiguità
            except Exception:
                return Response({'detail': f"Formato cella non valido: {c}"}, status=status.HTTP_400_BAD_REQUEST)

            n, _ = Assignment.objects.filter(
                plan=plan,
                profession_id=prof_id,
                date=day,
            ).delete()
            deleted += n

        return Response({'deleted': deleted}, status=status.HTTP_200_OK)


    @action(detail=True, methods=['post'])
    def set_note(self, request, pk=None):
        """
        Imposta/aggiorna/azzera la nota su una cella (deve esistere un Assignment).
        Body: { profession_id: int, date: 'YYYY-MM-DD', note: '...' }
        """
        plan = self.get_object()
        try:
            prof_id = int(request.data.get('profession_id'))
            day = dt.date.fromisoformat(request.data.get('date'))
        except Exception:
            return Response({'detail': 'profession_id/date non validi'}, status=status.HTTP_400_BAD_REQUEST)

        note = (request.data.get('note') or '').strip()

        try:
            a = Assignment.objects.get(plan=plan, profession_id=prof_id, date=day)
        except Assignment.DoesNotExist:
            return Response({'detail': 'Nessuna assegnazione su questa cella. Assegna un lavoratore prima di aggiungere una nota.'},
                            status=status.HTTP_400_BAD_REQUEST)

        a.notes = note  # può essere stringa vuota per rimuovere la nota
        a.save(update_fields=['notes'])

        return Response({'ok': True, 'notes': a.notes, 'has_note': bool(a.notes)}, status=status.HTTP_200_OK)


    @action(detail=True, methods=['get'])
    def export_xlsx(self, request, pk=None):
        plan = self.get_object()
        days = calendar.monthrange(plan.year, plan.month)[1]

        # Prepara workbook/worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = f"{plan.month:02d}-{plan.year}"

        # Stili base
        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wrap = Alignment(wrap_text=True)
        header_fill = PatternFill("solid", fgColor="E3F2FD")     # azzurrino header
        red_font = Font(bold=True, color="FFD32F2F")

        # Header: A1 "Professione", poi 1..N con giorno + abbreviazione
        ws.cell(row=1, column=1, value="Professione").font = bold
        ws.cell(row=1, column=1).alignment = center
        ws.cell(row=1, column=1).fill = header_fill

        it_hdays = italy_holidays(plan.year)
        it_weekdays = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]
        # mappa indice colonna -> (giorno, data)
        col_for_day = {}
        for d in range(1, days + 1):
            col = d + 1
            the_date = dt.date(plan.year, plan.month, d)
            wd = it_weekdays[the_date.weekday()]

            c = ws.cell(row=1, column=col, value=f"{d}\n{wd}")
            c.alignment = center
            c.fill = header_fill

            # sabato/dom/holiday -> numeri rossi
            if the_date.weekday() in (5, 6) or the_date in it_hdays:
                c.font = red_font
            else:
                c.font = bold

            col_for_day[col] = the_date

        # -------- Layout righe: usa PlanRow se presenti, altrimenti Profession --------
        prof_by_name = {p.name.strip().casefold(): p for p in Profession.objects.all()}
        use_plan_rows = plan.rows.exists()
        if use_plan_rows:
            layout = []
            for r in plan.rows.all().order_by('order'):
                if r.is_spacer:
                    layout.append({'label': '', 'profession': None, 'spacer': True})
                else:
                    layout.append({
                        'label': r.duty,
                        'profession': prof_by_name.get((r.duty or '').strip().casefold()),
                        'spacer': False
                    })
        else:
            layout = [{'label': p.name, 'profession': p, 'spacer': False}
                      for p in Profession.objects.order_by('name')]

        # -------- Precarica assignments --------
        ass = (Assignment.objects
               .filter(plan=plan)
               .select_related('employee', 'shift_type', 'profession'))
        idx = {(a.profession_id, a.date): a for a in ass}

        # -------- Scrittura righe --------
        row = 2
        for item in layout:
            # Colonna A: etichetta del template (senza suffisso .N; vuota per spacer)
            label_full = item.get('label') or ''
            base_label, _ = _split_slot(label_full)  # <- usa helper già definito
            ws.cell(row=row, column=1, value=base_label if not item.get('spacer') else '')

            # Celle giorno per giorno
            for d in range(1, days+1):
                col = d + 1
                day = col_for_day[col]

                # Spacer o mansione non mappata a Profession -> lascia vuoto
                if item['spacer'] or not item['profession']:
                    continue

                p = item['profession']
                a = idx.get((p.id, day))
                if not a:
                    continue

                name = a.employee.full_name()
                shift = f" ({a.shift_type.label})" if a.shift_type else ""
                cell = ws.cell(row=row, column=col, value=f"{name}{shift}")
                cell.alignment = wrap
                if a.notes:
                    try:
                        cell.comment = Comment(a.notes, "ShiftPlanner")
                    except Exception:
                        pass
            row += 1


        # layout
        ws.freeze_panes = "B2"
        ws.column_dimensions["A"].width = 40
        # larghezza base per giorni
        for col in range(2, days+2):
            ws.column_dimensions[get_column_letter(col)].width = 22

        # Output
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        fname = f"plan_{plan.year}_{plan.month:02d}.xlsx"
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp

@login_required
def home(request):
    return render(request, 'home.html')

@login_required
def functions_hub(request):
    return render(request, "home.html")

@login_required
def calendar_notes(request):
    plans = Plan.objects.order_by('-year','-month')[:12]
    return render(request, 'calendar_notes.html', {'plans': plans})

@login_required
def profile(request):
    user = request.user
    return render(request, 'profile.html', {
        'user_obj': user,
    })

@login_required
def monthly_plan(request, pk: int):
    plan = get_object_or_404(Plan, pk=pk)
    employees = Employee.objects.order_by('last_name','first_name')
    shifts = ShiftType.objects.order_by('id')
    use_plan_rows = plan.rows.exists()
    return render(request, 'monthly_plan.html', {
        'plan': plan,
        'use_plan_rows': use_plan_rows,
        'employees': employees,
        'shifts': shifts,
    })

def logout_view(request):
    logout(request)
    return redirect('login')

def _is_staff_or_superuser(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)

@login_required
@user_passes_test(_is_staff_or_superuser)
def plans_area(request):
    plans = Plan.objects.order_by("-year","-month","name")
    return render(request, "plans_area.html", {"plans": plans})

@login_required
@user_passes_test(_is_staff_or_superuser)
def templates_area(request):
    tpls = Template.objects.order_by("name")
    return render(request, "templates_area.html", {"templates": tpls})

@login_required
def employees_directory(request):
    """
    Elenco dipendenti con filtri lato client (azienda + nome/cognome).
    """
    employees = (
        Employee.objects
        .select_related("company")
        .order_by("last_name", "first_name")
    )
    companies = Company.objects.order_by("name")

    return render(request, "employees_directory.html", {
        "employees": employees,
        "companies": companies,
    })

@login_required
@user_passes_test(_is_staff_or_superuser)
def plan_create(request):
    if request.method == "POST":
        form = PlanCreateForm(request.POST)
        if form.is_valid():
            month = int(form.cleaned_data["month"])
            year = int(form.cleaned_data["year"])

            existing = Plan.objects.filter(month=month, year=year).first()
            if existing:
                messages.info(request, "Esiste già un piano per questo mese/anno. Apertura del piano esistente.")
                return redirect("monthly_plan", pk=existing.pk)

            plan = form.save(commit=False)
            plan.created_by = request.user
            plan.save()

            # --- Clona righe dal template, se selezionato ---
            template = form.cleaned_data.get("template")
            if template:
                from collections import defaultdict
                from .models import TemplateRow, PlanRow, Profession

                occ = defaultdict(int)  # contatore per base NEL SOLO PIANO
                rows_to_create = []

                for r in template.rows.all().order_by('order'):
                    if r.is_spacer:
                        rows_to_create.append(PlanRow(
                            plan=plan, order=r.order, duty="", is_spacer=True, notes=r.notes
                        ))
                        continue

                    raw = (r.duty or "").strip()
                    if not raw:
                        rows_to_create.append(PlanRow(
                            plan=plan, order=r.order, duty="", is_spacer=False, notes=r.notes
                        ))
                        continue

                    # split suffisso esplicito
                    parts = raw.rsplit(".", 1)
                    if len(parts) == 2 and parts[1].isdigit():
                        base, num = parts[0].strip(), int(parts[1])
                        duty_full = f"{base}.{num}"
                        occ[base] = max(occ[base], num)  # allinea contatore
                    else:
                        base = raw
                        occ[base] += 1
                        duty_full = f"{base}.{occ[base]}"

                    Profession.objects.get_or_create(name=duty_full)  # accredita se manca

                    rows_to_create.append(PlanRow(
                        plan=plan, order=r.order, duty=duty_full, is_spacer=False, notes=r.notes
                    ))

                if rows_to_create:
                    PlanRow.objects.bulk_create(rows_to_create)

            messages.success(request, "Piano creato correttamente.")
            return redirect("monthly_plan", pk=plan.pk)
    else:
        form = PlanCreateForm()

    return render(request, "plan_create.html", {"form": form})


def _parse_rows(text: str):
    """Ritorna un elenco di TemplateRow da testo multilinea."""
    lines = text.replace("\r\n", "\n").split("\n")
    rows = []
    order = 1
    for ln in lines:
        stripped = ln.strip()
        if stripped == "":
            rows.append(TemplateRow(order=order, duty="", is_spacer=True))
        else:
            rows.append(TemplateRow(order=order, duty=stripped, is_spacer=False))
        order += 1
    # elimina eventuali spazi finali consecutivi
    while rows and rows[-1].is_spacer:
        rows.pop()
    return rows

@login_required
@user_passes_test(_is_staff_or_superuser)
@transaction.atomic
def template_create(request):
    """Form per creare un nuovo template di piano turni + accredito Profession."""
    if request.method == "POST":
        form = TemplateCreateForm(request.POST)
        if form.is_valid():
            template = form.save()
            rows = _parse_rows(form.cleaned_data["rows_text"])
            for r in rows:
                r.template = template

            # --- ACCREDITO PROFESSION in base alle righe inserite ---
            from .models import Profession, TemplateRow  # import locale per chiarezza
            base_counters = defaultdict(int)  # occorrenze per base nel SOLO template

            for r in rows:
                if r.is_spacer:
                    continue
                base, num = _split_slot(r.duty)
                if not base:
                    continue

                if num > 0:
                    # suffisso esplicito: crea esattamente base.num se manca
                    desired = f"{base}.{num}"
                    Profession.objects.get_or_create(name=desired)
                else:
                    # senza suffisso: usa il prossimo disponibile considerando DB + occorrenze nel template
                    base_counters[base] += 1
                    idx_in_template = base_counters[base]
                    start = _existing_max_suffix(base)       # max già presente in DB
                    target_num = start + idx_in_template     # continua la sequenza
                    desired = f"{base}.{target_num}"
                    Profession.objects.get_or_create(name=desired)

            # --- Salvataggio righe template così come inserite dall’utente ---
            if rows:
                TemplateRow.objects.bulk_create(rows)

            messages.success(request, "Template creato correttamente.")
            return redirect("template_detail", pk=template.pk)
    else:
        form = TemplateCreateForm()
    return render(request, "template_create.html", {"form": form})


@login_required
@user_passes_test(_is_staff_or_superuser)
def template_detail(request, pk):
    """Visualizza un template e le sue righe."""
    template = get_object_or_404(Template.objects.prefetch_related("rows"), pk=pk)
    return render(request, "template_detail.html", {"template": template})


@login_required
@user_passes_test(_is_staff_or_superuser)
@require_POST
@transaction.atomic
def template_insert_row(request, pk: int):
    import json
    try:
        data = json.loads(request.body.decode('utf-8') or "{}")
    except Exception:
        return JsonResponse({"detail": "JSON non valido"}, status=400)

    pos = (data.get('position') or 'after').lower()
    tr_id = data.get('template_row_id')
    duty = (data.get('duty') or '').strip() or None
    base = (data.get('base') or '').strip()

    if pos not in ('after', 'before'):
        return JsonResponse({"detail": "position deve essere 'after' o 'before'."}, status=400)
    if not tr_id:
        return JsonResponse({"detail": "template_row_id obbligatorio."}, status=400)

    try:
        template = Template.objects.prefetch_related('rows').get(pk=pk)
    except Template.DoesNotExist:
        return JsonResponse({"detail": "Template non trovato."}, status=404)

    try:
        ref = TemplateRow.objects.get(pk=int(tr_id), template=template)
    except TemplateRow.DoesNotExist:
        return JsonResponse({"detail": "template_row_id non appartiene al template."}, status=404)

    insert_order = ref.order + (1 if pos == 'after' else 0)
    if pos == 'before':
        insert_order = max(1, ref.order)

    # normalizza/risolve duty -> Profession
    if not duty:
        if not base:
            return JsonResponse({"detail": "Serve 'base' se non passi 'duty'."}, status=400)
        prof = _ensure_profession_for(base, None)
        duty = prof.name
    else:
        b, num = _split_slot(duty)
        if not b:
            return JsonResponse({"detail": "duty non valido."}, status=400)
        prof = _ensure_profession_for(b, num or None)
        duty = prof.name  # normalizza

    # --- BUMP -> INSERT -> NORMALIZE sul Template ---
    BUMP = 1000
    TemplateRow.objects.filter(template=template, order__gte=insert_order).update(order=F('order') + BUMP)
    new_tr = TemplateRow.objects.create(template=template, order=insert_order, duty=duty, is_spacer=False)
    _normalize_template_orders(template)

    # Propaga ai Plan con lo stesso schema
    _propagate_insert_to_plans(template, insert_order, duty)

    plan_ids = list(Plan.objects.filter(template=template).values_list('id', flat=True))
    return JsonResponse({
        "ok": True,
        "template_id": template.id,
        "template_row_id": new_tr.id,
        "insert_order": insert_order,
        "duty": duty,
        "profession_id": prof.id,
        "propagated_plans": plan_ids
    }, status=201)

@login_required
@user_passes_test(_is_staff_or_superuser)
@require_POST
@transaction.atomic
def template_delete_row(request, pk: int):
    """
    Elimina una riga del Template identificandola per 'order' su TUTTI i piani che usano quel template.
    Blocca se una delle righe mappate (stesso 'order' nel piano) ha assegnazioni.
    """
    import json
    from django.db.models import Count
    from .models import Template, TemplateRow, Plan, PlanRow, Assignment, Profession

    try:
        data = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return JsonResponse({"detail": "JSON non valido"}, status=400)

    tr_id = data.get("template_row_id")
    if not tr_id:
        return JsonResponse({"detail": "template_row_id obbligatorio."}, status=400)

    # Template + riga
    try:
        template = Template.objects.prefetch_related("rows").get(pk=pk)
    except Template.DoesNotExist:
        return JsonResponse({"detail": "Template non trovato."}, status=404)

    try:
        tr = TemplateRow.objects.get(pk=int(tr_id), template=template)
    except TemplateRow.DoesNotExist:
        return JsonResponse({"detail": "Riga non trovata nel template."}, status=404)

    target_order = tr.order  # chiave di mappatura tra TemplateRow e PlanRow

    # Se spacer: sempre eliminabile. Si propaga per 'order' nei piani.
    if tr.is_spacer:
        tr.delete()
        # elimina PlanRow alla stessa posizione nei piani
        for plan in Plan.objects.filter(template=template):
            pr = PlanRow.objects.filter(plan=plan, order=target_order).first()
            if pr:
                pr.delete()
                _normalize_plan_orders(plan)
        _normalize_template_orders(template)
        return JsonResponse({"ok": True, "deleted_spacer": True, "order": target_order}, status=200)

    # Riga con duty: blocca se QUALSIASI piano ha assegnazioni su quella riga (per order)
    blocking = []
    affected_plan_ids = []
    plans = list(Plan.objects.filter(template=template))
    for plan in plans:
        pr = PlanRow.objects.filter(plan=plan, order=target_order).first()
        if not pr:
            continue  # piano desincronizzato: lo saltiamo, sarà normalizzato dopo
        affected_plan_ids.append(plan.id)

        duty_name = (pr.duty or "").strip()
        if not duty_name:
            continue  # riga vuota nel piano → non blocca

        # check assegnazioni sulla profession del piano
        has_ass = Assignment.objects.filter(
            plan=plan, profession__name=duty_name
        ).select_related("employee", "shift_type").exists()

        if has_ass:
            # raccogli fino a 20 dettagli per feedback
            ass_details = list(
                Assignment.objects.filter(plan=plan, profession__name=duty_name)
                .select_related("employee", "shift_type")
                .order_by("date")[:20]
            )
            for a in ass_details:
                blocking.append({
                    "piano": plan.name,
                    "data": a.date.strftime("%d/%m/%Y"),
                    "dipendente": a.employee.full_name() if a.employee_id else "",
                    "turno": (a.shift_type.label if a.shift_type_id else ""),
                    "mansione": duty_name,
                })

    if blocking:
        # blocca e spiega cosa rimuovere prima
        return JsonResponse({
            "ok": False,
            "detail": (
                "Impossibile eliminare la riga: esistono assegnazioni attive sulla mansione "
                f"(posizione #{target_order}) in uno o più piani. Rimuovi prima tali assegnazioni."
            ),
            "assegnazioni": blocking
        }, status=409)

    # Nessun blocco: elimina dal template e dai piani alla stessa 'order'
    deleted_duty = (tr.duty or "").strip()
    tr.delete()

    for plan in plans:
        pr = PlanRow.objects.filter(plan=plan, order=target_order).first()
        if pr:
            pr.delete()
            _normalize_plan_orders(plan)

    _normalize_template_orders(template)

    return JsonResponse({
        "ok": True,
        "deleted_duty": deleted_duty,
        "order": target_order,
        "propagated_plans": affected_plan_ids,
    }, status=200)


@login_required
def privacy(request):
    return render(request, "privacy/privacy.html")

@login_required
def analytics_overview(request):
    # aziende con almeno una assegnazione
    company_ids = (Assignment.objects
                   .values_list("employee__company_id", flat=True)
                   .distinct())
    companies = Company.objects.filter(id__in=company_ids).order_by("name")

    # default periodo = ultimi 30 giorni
    start, end = _range_from_preset(None)
    ctx = {
        "companies": companies,
        "default_start": start.isoformat(),
        "default_end": end.isoformat(),
    }
    return render(request, "analytics_overview.html", ctx)

@login_required
def analytics_summary(request):
    """
    GET /analytics/summary/?company_id=ID&start=YYYY-MM-DD&end=YYYY-MM-DD&preset=m1|q1|h1|y1
    JSON per grafici e tabelle.
    """
    company_id = request.GET.get("company_id")
    preset = request.GET.get("preset")
    start_s = request.GET.get("start")
    end_s = request.GET.get("end")

    if preset:
        start, end = _range_from_preset(preset)
    else:
        start = parse_date(start_s) if start_s else None
        end = parse_date(end_s) if end_s else None
        if not (start and end):
            start, end = _range_from_preset(None)

    qs = Assignment.objects.select_related("employee","profession","shift_type","plan") \
                           .filter(date__range=(start, end))
    if company_id:
        qs = qs.filter(employee__company_id=company_id)  # cambia in plan__company_id se l'azienda è sul Plan

    # --- Esclusioni per "assenze/permessi" ---
    EXCLUDE_WORDS = [
        "Formazione",
        "Ferie",
        "Permessi",
        "Congedo",
        "Congedo Matrimoniale",
        "Permesso 104",
        "Permesso Sindacale",
        "Malattia",
        "Assenza",
        "Sciopero",
    ]
    # filtro OR su profession__name icontains
    q_or = Q()
    for kw in EXCLUDE_WORDS:
        q_or |= Q(profession__name__icontains=kw)

    qs_excluded = qs.filter(q_or)
    qs_turno = qs.exclude(q_or)

    # KPI
    total_all = qs.count()
    total_turno = qs_turno.count()

    # dipendenti assegnati nel periodo (distinti)
    employees_assigned = qs.values("employee_id").distinct().count()

    # dipendenti totali registrati nel sistema
    if company_id:
        employees_total = Employee.objects.filter(company_id=company_id).count()
        # Se l'azienda è su Plan e NON su Employee, usa:
        # employees_total = Employee.objects.filter(assignment__plan__company_id=company_id).distinct().count()
    else:
        employees_total = Employee.objects.count()

    # breakdown esclusi per parola
    excluded_breakdown = []
    excluded_total = 0
    for kw in EXCLUDE_WORDS:
        n = qs.filter(profession__name__icontains=kw).count()
        excluded_breakdown.append({"label": kw, "count": n})
        excluded_total += n

    # Tabelle già presenti
    per_shift = (qs.values("shift_type__id","shift_type__code","shift_type__label")
                   .annotate(assignments=Count("id"),
                             employees=Count("employee_id", distinct=True))
                   .order_by("shift_type__id"))

    preposto = (qs.filter(profession__name__icontains="preposto")
                  .values("profession__name")
                  .annotate(assignments=Count("id"),
                            employees=Count("employee_id", distinct=True))
                  .order_by("-assignments"))

    chart_shift = {
        "labels": [r["shift_type__label"] or r["shift_type__code"] or "—" for r in per_shift],
        "datasets": [
            {"label": "Assegnazioni", "data": [r["assignments"] for r in per_shift]},
            {"label": "Dipendenti distinti", "data": [r["employees"] for r in per_shift]},
        ]
    }
    chart_preposto = {
        "labels": [r["profession__name"] for r in preposto] or ["Nessuna"],
        "datasets": [
            {"label": "Assegnazioni", "data": [r["assignments"] for r in preposto] or [0]},
            {"label": "Dipendenti distinti", "data": [r["employees"] for r in preposto] or [0]},
        ]
    }

    return JsonResponse({
        "scope": {"start": start.isoformat(), "end": end.isoformat(), "company_id": company_id},
        "kpi": {
            "assignments_all": total_all,
            "assignments_turno": total_turno,
            "employees_assigned": employees_assigned,
            "employees_total": employees_total,
            "excluded_total": excluded_total,
            "excluded_breakdown": excluded_breakdown,
        },
        "tables": {"per_shift": list(per_shift), "preposto": list(preposto)},
        "charts": {"per_shift": chart_shift, "preposto": chart_preposto},
    })
